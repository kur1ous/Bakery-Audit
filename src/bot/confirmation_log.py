from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook, load_workbook

from .models import BetExtraction

HEADERS = [
    "date",
    "team",
    "against",
    "odds",
    "stake",
    "return",
    "num stake",
    "num return",
]

CURRENCY_USD = "USD"
CURRENCY_CAD = "CAD"


@dataclass(frozen=True)
class ConfirmationLogContext:
    message_id: int
    channel_id: int
    guild_id: int | None
    invoker_user_id: int
    has_hedge_pair: bool


class ConfirmationLogger(Protocol):
    def log_batch(
        self,
        context: ConfirmationLogContext,
        extractions: list[BetExtraction],
        bet_currencies: list[str] | None = None,
        usd_to_cad_rate: float = 1.36,
    ) -> int:
        ...


class ExcelConfirmationLogger:
    def __init__(self, file_path: str, sheet_name: str = "confirmed_bets") -> None:
        self.file_path = file_path
        self.sheet_name = sheet_name
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        path = Path(self.file_path)
        if path.parent and str(path.parent) not in ("", "."):
            path.parent.mkdir(parents=True, exist_ok=True)

        if path.exists():
            wb = load_workbook(path)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.create_sheet(self.sheet_name)
            if ws.max_row == 0:
                ws.append(HEADERS)
            elif ws.max_row >= 1:
                existing_headers = [ws.cell(row=1, column=idx + 1).value for idx in range(len(HEADERS))]
                if existing_headers != HEADERS:
                    for idx, header in enumerate(HEADERS, start=1):
                        ws.cell(row=1, column=idx, value=header)
            wb.save(path)
            wb.close()
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        ws.append(HEADERS)
        wb.save(path)
        wb.close()

    def log_batch(
        self,
        context: ConfirmationLogContext,
        extractions: list[BetExtraction],
        bet_currencies: list[str] | None = None,
        usd_to_cad_rate: float = 1.36,
    ) -> int:
        rows = [
            _row(extraction, _currency_for_index(bet_currencies or [], idx), usd_to_cad_rate)
            for idx, extraction in enumerate(extractions)
        ]

        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]
        for row in rows:
            ws.append(row)
        wb.save(self.file_path)
        wb.close()
        return len(rows)


class GoogleSheetsConfirmationLogger:
    def __init__(self, spreadsheet_id: str, credentials_json_path: str, worksheet_name: str = "confirmed_bets") -> None:
        self.spreadsheet_id = spreadsheet_id
        self.credentials_json_path = credentials_json_path
        self.worksheet_name = worksheet_name

    def _worksheet(self):
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        credentials = Credentials.from_service_account_file(self.credentials_json_path, scopes=scopes)
        client = gspread.authorize(credentials)

        try:
            spreadsheet = client.open_by_key(self.spreadsheet_id)
        except PermissionError as exc:
            raise RuntimeError(
                "Google Sheets access denied (403). Share the sheet with your service-account email as Editor."
            ) from exc

        try:
            worksheet = spreadsheet.worksheet(self.worksheet_name)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=self.worksheet_name, rows=1000, cols=32)
            worksheet.append_row(HEADERS)

        return worksheet

    def log_batch(
        self,
        context: ConfirmationLogContext,
        extractions: list[BetExtraction],
        bet_currencies: list[str] | None = None,
        usd_to_cad_rate: float = 1.36,
    ) -> int:
        rows = [
            _row(extraction, _currency_for_index(bet_currencies or [], idx), usd_to_cad_rate)
            for idx, extraction in enumerate(extractions)
        ]

        worksheet = self._worksheet()
        header_values = worksheet.get("1:1")[0] if worksheet.get("1:1") else []
        if header_values[: len(HEADERS)] != HEADERS:
            worksheet.clear()
            worksheet.append_row(HEADERS)

        worksheet.append_rows(rows, value_input_option="RAW")
        return len(rows)


def create_confirmation_logger(
    *,
    backend: str,
    excel_path: str,
    google_sheet_id: str,
    google_credentials_json_path: str,
    google_worksheet_name: str,
) -> ConfirmationLogger:
    normalized = backend.strip().lower()
    if normalized == "excel":
        return ExcelConfirmationLogger(file_path=excel_path, sheet_name="confirmed_bets")

    if normalized == "google_sheets":
        if not google_sheet_id:
            raise RuntimeError("CONFIRM_GOOGLE_SHEET_ID is required for google_sheets logging backend")
        if not google_credentials_json_path:
            raise RuntimeError("CONFIRM_GOOGLE_CREDENTIALS_JSON is required for google_sheets logging backend")

        return GoogleSheetsConfirmationLogger(
            spreadsheet_id=google_sheet_id,
            credentials_json_path=google_credentials_json_path,
            worksheet_name=google_worksheet_name or "confirmed_bets",
        )

    raise RuntimeError("CONFIRM_LOG_BACKEND must be 'excel' or 'google_sheets'")


def _row(extraction: BetExtraction, currency: str, usd_to_cad_rate: float) -> list[str | float]:
    # Internal policy: sheets are always logged in USD.
    stake_usd = _to_number(extraction.stake)
    return_usd = _to_number(extraction.return_amount)

    if currency == CURRENCY_CAD:
        # User selected CAD for this bet: convert displayed CAD back to USD for logging.
        stake_usd = _cad_to_usd(stake_usd, usd_to_cad_rate)
        return_usd = _cad_to_usd(return_usd, usd_to_cad_rate)

    stake_val = round(stake_usd, 2) if stake_usd is not None else ""
    return_val = round(return_usd, 2) if return_usd is not None else ""

    return [
        extraction.date,
        _to_team_code(extraction.team),
        _to_team_code(extraction.against),
        extraction.odds,
        stake_val,
        return_val,
        stake_val,
        return_val,
    ]


def _to_number(value: str) -> float | None:
    cleaned = (value or "").strip().replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _cad_to_usd(value: float | None, usd_to_cad_rate: float) -> float | None:
    if value is None:
        return None
    if usd_to_cad_rate <= 0:
        return value
    return value / usd_to_cad_rate


def _currency_for_index(currencies: list[str], idx: int) -> str:
    if 0 <= idx < len(currencies):
        c = (currencies[idx] or "").strip().upper()
        if c in (CURRENCY_USD, CURRENCY_CAD):
            return c
    return CURRENCY_USD


def _to_team_code(value: str) -> str:
    text = (value or "").strip().upper()
    if not text:
        return ""

    tokens = [t for t in re_split_non_alnum_(text) if t]
    if not tokens:
        return ""

    first = tokens[0]
    if len(first) == 3:
        return first

    if len(first) == 2 and len(tokens) > 1 and tokens[1]:
        return (first + tokens[1][0])[:3]

    if len(first) >= 3:
        return first[:3]

    compact = "".join(tokens)
    return (compact + "XXX")[:3]


def re_split_non_alnum_(text: str) -> list[str]:
    parts = []
    current = []
    for ch in text:
        if ch.isalnum():
            current.append(ch)
        else:
            if current:
                parts.append("".join(current))
                current = []
    if current:
        parts.append("".join(current))
    return parts