from __future__ import annotations

from openpyxl import load_workbook

from src.bot.confirmation_log import (
    ConfirmationLogContext,
    ExcelConfirmationLogger,
    create_confirmation_logger,
)
from src.bot.models import BetExtraction


def test_excel_log_batch_writes_one_row_per_bet_in_usd(tmp_path) -> None:
    file_path = tmp_path / "confirmed_bets.xlsx"
    logger = ExcelConfirmationLogger(str(file_path))

    context = ConfirmationLogContext(
        message_id=999,
        channel_id=123,
        guild_id=456,
        invoker_user_id=777,
        has_hedge_pair=True,
    )
    bets = [
        BetExtraction.model_validate(
            {
                "date": "2026-04-01",
                "team": "Miami Heat",
                "against": "Boston Celtics",
                "odds": "2.69",
                "stake": "80",
                "return": "283.41",
                "confidence": 0.9,
            }
        ),
        BetExtraction.model_validate(
            {
                "date": "2026-04-01",
                "team": "TOR Raptors",
                "against": "DET Pistons",
                "odds": "1.52",
                "stake": "75",
                "return": "180.0",
                "confidence": 0.9,
            }
        ),
    ]

    written = logger.log_batch(context, bets, ["USD", "USD"], 1.36)
    assert written == 2

    wb = load_workbook(file_path)
    ws = wb["confirmed_bets"]
    assert ws.max_row == 3

    headers = [ws.cell(row=1, column=col).value for col in range(1, 9)]
    assert headers == ["date", "team", "against", "odds", "stake", "return", "num stake", "num return"]

    first_row = [ws.cell(row=2, column=col).value for col in range(1, 9)]
    assert first_row == ["2026-04-01", "MIA", "BOS", "2.69", 80.0, 283.41, 80.0, 283.41]

    second_row = [ws.cell(row=3, column=col).value for col in range(1, 9)]
    assert second_row[1] == "TOR"
    assert second_row[2] == "DET"
    wb.close()


def test_excel_log_batch_converts_cad_to_usd_when_selected(tmp_path) -> None:
    file_path = tmp_path / "confirmed_bets.xlsx"
    logger = ExcelConfirmationLogger(str(file_path))

    context = ConfirmationLogContext(
        message_id=100,
        channel_id=10,
        guild_id=None,
        invoker_user_id=20,
        has_hedge_pair=False,
    )
    bets = [
        BetExtraction.model_validate(
            {
                "date": "2026-04-02",
                "team": "Detroit Pistons",
                "against": "Toronto Raptors",
                "odds": "1.64",
                "stake": "108.8",
                "return": "69.74",
            }
        )
    ]

    written = logger.log_batch(context, bets, ["CAD"], 1.36)
    assert written == 1

    wb = load_workbook(file_path)
    ws = wb["confirmed_bets"]
    row = [ws.cell(row=2, column=col).value for col in range(1, 9)]
    assert row[1] == "DET"
    assert row[2] == "TOR"
    assert row[4] == 80.0
    assert row[5] == 51.28
    assert row[6] == 80.0
    assert row[7] == 51.28
    wb.close()


def test_create_confirmation_logger_defaults_to_excel(tmp_path) -> None:
    logger = create_confirmation_logger(
        backend="excel",
        excel_path=str(tmp_path / "bets.xlsx"),
        google_sheet_id="",
        google_credentials_json_path="",
        google_worksheet_name="confirmed_bets",
    )
    assert isinstance(logger, ExcelConfirmationLogger)


def test_google_sheets_backend_requires_settings(tmp_path) -> None:
    try:
        create_confirmation_logger(
            backend="google_sheets",
            excel_path=str(tmp_path / "bets.xlsx"),
            google_sheet_id="",
            google_credentials_json_path="",
            google_worksheet_name="confirmed_bets",
        )
        assert False, "Expected RuntimeError"
    except RuntimeError as exc:
        assert "CONFIRM_GOOGLE_SHEET_ID" in str(exc)